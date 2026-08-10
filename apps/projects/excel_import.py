from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils.exceptions import InvalidFileException
from django.utils.translation import gettext_lazy as _

from .models import SubObject

IMPORT_HEADERS = [_("Object"), _("Pod object (optional)")]


class InvalidExcelStructureError(Exception):
    pass


def build_subobjects_import_template():
    """An .xlsx a user can fill in and re-upload to bulk-create the project's
    objects/pod objects. Column A is the object name; column B, left blank on
    an object's own row, only needs a pod object name — it inherits the object
    from the nearest row above that had one filled in (like a merged cell)."""
    wb = Workbook()
    ws = wb.active
    ws.append([str(h) for h in IMPORT_HEADERS])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.append(["Bino 1", ""])
    ws.append(["", "KPP"])
    ws.append(["", "Garaj"])
    ws.append(["Bino 2", ""])
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30
    return wb


def import_subobjects_from_excel(project, uploaded_file):
    """Parses an .xlsx built from build_subobjects_import_template() and creates
    the objects/pod objects it describes under `project`. Existing objects/pod
    objects with matching names are reused rather than duplicated. Returns
    (objects_created, pods_created)."""
    try:
        wb = load_workbook(uploaded_file, read_only=True, data_only=True)
    except (InvalidFileException, KeyError, OSError) as exc:
        raise InvalidExcelStructureError(_("Could not read the Excel file.")) from exc

    ws = wb.active
    existing_objects = {
        so.name: so for so in SubObject.objects.filter(project=project, parent__isnull=True)
    }
    objects_created = 0
    pods_created = 0
    current_object = None

    for row_index, row in enumerate(ws.iter_rows(min_row=2, max_col=2, values_only=True), start=2):
        object_name = (str(row[0]).strip() if row and row[0] is not None else "")
        pod_name = (str(row[1]).strip() if row and len(row) > 1 and row[1] is not None else "")

        if object_name:
            current_object = existing_objects.get(object_name)
            if current_object is None:
                current_object = SubObject.objects.create(project=project, name=object_name)
                existing_objects[object_name] = current_object
                objects_created += 1

        if pod_name:
            if current_object is None:
                raise InvalidExcelStructureError(
                    _("Row %(row)d has a pod object but no object above it.") % {"row": row_index}
                )
            if not SubObject.objects.filter(
                project=project, parent=current_object, name=pod_name,
            ).exists():
                SubObject.objects.create(project=project, parent=current_object, name=pod_name)
                pods_created += 1

    return objects_created, pods_created
