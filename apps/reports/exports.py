from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def _make_workbook(headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(row)
    for i, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(str(header)) + 4)
    return wb


def build_projects_workbook(projects):
    headers = ["Name", "Client", "Stage", "Status", "Budget", "Paid", "Deadline"]
    rows = [
        [p.name, p.client_name, p.get_stage_display(), p.get_status_display(), p.budget, p.paid_amount, p.deadline]
        for p in projects
    ]
    return _make_workbook(headers, rows)


def build_tasks_workbook(tasks):
    headers = ["Title", "Project", "Assignee", "Status", "Priority", "Deadline"]
    rows = [
        [t.title, t.project.name, str(t.assignee) if t.assignee else "", t.get_status_display(), t.get_priority_display(), t.deadline]
        for t in tasks
    ]
    return _make_workbook(headers, rows)


def build_finance_workbook(records):
    headers = ["Project", "Type", "Amount", "Currency", "Date", "Status", "Description"]
    rows = [
        [r.project.name, r.get_type_display(), r.amount, r.currency, r.date, r.get_status_display(), r.description]
        for r in records
    ]
    return _make_workbook(headers, rows)


def build_employees_workbook(rows_data):
    headers = ["Employee", "Tasks total", "Tasks completed", "Contract amount", "Paid", "Balance"]
    rows = [
        [d["name"], d["tasks_total"], d["tasks_completed"], d["contract_amount"], d["paid"], d["balance"]]
        for d in rows_data
    ]
    return _make_workbook(headers, rows)
